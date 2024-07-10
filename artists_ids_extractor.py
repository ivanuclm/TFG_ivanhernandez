import openpyxl
from openpyxl import load_workbook
import pickle
import pandas as pd
from bs4 import BeautifulSoup
import re

import requests




def get_artist_ids_excel(filepath, sheet_name='Hoja1'):
    
    wb = load_workbook(filename = filepath, data_only=True)
    ws = wb[sheet_name]
    
    artistas = []
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
        for cell in row:
            artistas.append(cell.value)
    
    #remove duplicates but keep order
    artistas = list(dict.fromkeys(artistas))
    return artistas

def get_artist_ids_html(filepath):
    # Assuming `html_content` is the HTML content of your file
    with open(filepath, 'r', encoding='utf-8') as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, 'html.parser')

    # Find all <a> tags within <tr> tags
    artist_links = soup.find_all('tr')
    artistas = []
    for link in artist_links:
        a_tag = link.find('a', href=True)
        if a_tag:
            href = a_tag['href']
            # Extract the artist ID using regex
            match = re.search(r'artist/([0-9A-Za-z]+)', href)
            if match:
                artist_id = match.group(1)
                artistas.append(artist_id)

    artistas = list(dict.fromkeys(artistas))
    return artistas

def get_artist_ids_online(url):
    response = requests.get(url)
    if(response.status_code != 200):
        print('Error al obtener la página con el listado de artistas.')
        return None
    
    if(response.text == ''):
        print('La página con el listado de artistasestá vacía')
        return None
    

    soup = BeautifulSoup(response.text, 'html.parser')


    # Find all <a> tags within <tr> tags
    artist_links = soup.find_all('tr')
    artistas = []
    for link in artist_links:
        a_tag = link.find('a', href=True)
        if a_tag:
            href = a_tag['href']
            # Extract the artist ID using regex
            match = re.search(r'artist/([0-9A-Za-z]+)', href)
            if match:
                artist_id = match.group(1)
                artistas.append(artist_id)

    artistas = list(dict.fromkeys(artistas))
    return artistas

    


if __name__=='__main__':
    
    # filepath = 'input/excel/Top_Artistas_07052024.xlsm'
    # top_artist_ids = get_artist_ids_excel(filepath)

    # filepath = 'input/artists_listeners.html'
    # top_artist_ids = get_artist_ids_html(filepath)

    url = 'https://kworb.net/spotify/listeners.html'
    top_artist_ids = get_artist_ids_online(url)

    with open('input/top_artist_ids.pkl', 'wb') as f:
        pickle.dump(top_artist_ids, f)
    
    df = pd.DataFrame(top_artist_ids)
    df.to_csv('input/top_artist_ids.csv', index=False)
    print(f'Se ha extraído el ID de {len(top_artist_ids)} artistas.')
    
    